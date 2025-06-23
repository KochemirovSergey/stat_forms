import pandas as pd
import openpyxl
from pathlib import Path
from typing import List, Dict
import os
import re
from openpyxl.utils import get_column_letter

def get_excel_sheets(file_path: str) -> List[str]:
    """
    Получает список всех листов в Excel файле.
    
    Args:
        file_path (str): Путь к Excel файлу
        
    Returns:
        List[str]: Список названий листов
        
    Raises:
        FileNotFoundError: Если файл не найден
        ValueError: Если файл не является Excel файлом
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")
        
    if not file_path.suffix in ['.xlsx', '.xls']:
        raise ValueError(f"Файл {file_path} не является Excel файлом")
    
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    
    return sheet_names

def filter_section_sheets(sheets: List[str]) -> List[str]:
    """
    Фильтрует листы, оставляя только те, которые начинаются с 'Раздел'.
    
    Args:
        sheets (List[str]): Список названий листов
        
    Returns:
        List[str]: Отфильтрованный список листов
    """
    return [sheet for sheet in sheets if sheet.startswith('Раздел')]

def process_merged_cells(ws):
    """
    Обрабатывает объединенные ячейки, копируя значение в каждую ячейку диапазона.
    
    Args:
        ws: Рабочий лист openpyxl
    Returns:
        Dict: Словарь с значениями для объединенных ячеек
    """
    merged_values = {}
    
    if not ws.merged_cells:
        return merged_values
        
    merged_ranges = list(ws.merged_cells.ranges)
    if not merged_ranges:
        return merged_values
        
    for merged_range in merged_ranges:
        try:
            # Получаем значение из первой ячейки объединенного диапазона
            first_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if first_cell is None:
                continue
                
            value = first_cell.value
            
            # Сохраняем значения для каждой ячейки в диапазоне
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = value
            
        except Exception as e:
            print(f"Ошибка при обработке объединенного диапазона: {str(e)}")
            continue
            
    return merged_values

def clean_value(value) -> any:
    """
    Очищает значение от переносов строк.
    
    Args:
        value: Значение для очистки
        
    Returns:
        any: Очищенное значение
    """
    if isinstance(value, str):
        return value.replace('\n', ' ').replace('\r', ' ').strip()
    return value

def save_sheets_to_single_folder(excel_path: str):
    """
    Основная функция для обработки Excel файла и сохранения всех листов в одну папку.
    
    Args:
        excel_path (str): Путь к Excel файлу
    """
    try:
        # Получаем все листы
        print("Получаем список листов...")
        sheets = get_excel_sheets(excel_path)
        print(f"Найдено листов: {len(sheets)}")
        
        # Фильтруем только листы с разделами
        section_sheets = filter_section_sheets(sheets)
        print(f"Отфильтровано листов с разделами: {len(section_sheets)}")
        if not section_sheets:
            raise ValueError("Не найдено листов, начинающихся с 'Раздел'")
        
        # Создаем папку для сохранения всех листов
        excel_name = Path(excel_path).stem
        output_dir = Path(excel_path).parent / excel_name
        output_dir.mkdir(exist_ok=True, parents=True)
        print(f"Создана папка для сохранения: {output_dir}")
        
        # Загружаем исходный Excel файл
        print("Загружаем исходный файл...")
        source_wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Обрабатываем каждый лист
        for sheet in section_sheets:
            print(f"\nОбработка листа: {sheet}")
            
            # Получаем исходный лист
            source_ws = source_wb[sheet]
            print(f"Размер листа: {source_ws.max_row}x{source_ws.max_column}")
            
            # Получаем значения объединенных ячеек
            merged_values = {}
            if source_ws.merged_cells:
                print(f"Найдено объединенных диапазонов: {len(source_ws.merged_cells.ranges)}")
                merged_values = process_merged_cells(source_ws)
            
            # Создаем новый файл для листа
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            
            # Копируем данные
            for row in source_ws.rows:
                for cell in row:
                    if cell is None:
                        continue
                    
                    # Проверяем, есть ли значение в объединенных ячейках
                    value = merged_values.get((cell.row, cell.column), cell.value)
                    # Очищаем значение от переносов строк
                    value = clean_value(value)
                    
                    new_cell = new_ws.cell(
                        row=cell.row,
                        column=cell.column,
                        value=value
                    )
            
            # Копируем размеры столбцов
            for col in range(1, source_ws.max_column + 1):
                letter = get_column_letter(col)
                if letter in source_ws.column_dimensions:
                    new_ws.column_dimensions[letter].width = source_ws.column_dimensions[letter].width
            
            # Сохраняем результат
            output_path = output_dir / f"{sheet}.xlsx"
            print(f"Сохраняем в: {output_path}")
            new_wb.save(str(output_path))
            new_wb.close()
        
        source_wb.close()
        print("\nВсе листы успешно обработаны")
        
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise

if __name__ == "__main__":
    # Пример использования
    file_path = "/Users/sergejkocemirov/stat_forms/Таблицы_исходники/2016.xlsx"
    try:
        save_sheets_to_single_folder(file_path)
        print("Листы успешно сохранены в общую папку")
    except Exception as e:
        print(f"Ошибка: {str(e)}") 