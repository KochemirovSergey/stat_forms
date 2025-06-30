import os
import re
from pathlib import Path
from typing import List, Tuple, Dict
import traceback

# Поддержка как относительных, так и абсолютных импортов
try:
    from .excel_utils_single_folder import save_sheets_to_single_folder
except ImportError:
    from excel_utils_single_folder import save_sheets_to_single_folder

def extract_region_name(filename: str) -> str:
    """
    Извлекает название региона из имени файла.
    
    Args:
        filename (str): Имя файла в формате "Название региона (ГОУ) (город+село).xlsx"
        
    Returns:
        str: Название региона
        
    Example:
        "Алтайский край (ГОУ) (город+село).xlsx" -> "Алтайский край"
        "Кемеровская область – Кузбасс (ГОУ) (город+село).xlsx" -> "Кемеровская область – Кузбасс"
    """
    # Убираем расширение файла
    name_without_ext = Path(filename).stem
    
    # Ищем первые скобки и берем все до них
    match = re.match(r'^(.+?)\s*\(', name_without_ext)
    if match:
        return match.group(1).strip()
    
    # Если скобки не найдены, возвращаем имя без расширения
    return name_without_ext

def get_excel_files_from_folder(folder_path: str) -> List[str]:
    """
    Получает список всех Excel файлов из указанной папки.
    
    Args:
        folder_path (str): Путь к папке
        
    Returns:
        List[str]: Список путей к Excel файлам
    """
    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"Папка не найдена: {folder_path}")
    
    excel_files = []
    for file_path in folder.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in ['.xlsx', '.xls']:
            excel_files.append(str(file_path))
    
    return sorted(excel_files)

def save_sheets_to_region_folder(excel_path: str, region_name: str, base_folder: str) -> None:
    """
    Модифицированная версия функции save_sheets_to_single_folder для сохранения в папку региона.
    
    Args:
        excel_path (str): Путь к Excel файлу
        region_name (str): Название региона
        base_folder (str): Базовая папка (2024/)
    """
    # Поддержка как относительных, так и абсолютных импортов
    try:
        from .excel_utils_single_folder import (
            get_excel_sheets, filter_section_sheets, process_merged_cells,
            clean_value
        )
    except ImportError:
        from excel_utils_single_folder import (
            get_excel_sheets, filter_section_sheets, process_merged_cells,
            clean_value
        )
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    # Получаем все листы
    sheets = get_excel_sheets(excel_path)
    
    # Фильтруем только листы с разделами
    section_sheets = filter_section_sheets(sheets)
    if not section_sheets:
        raise ValueError("Не найдено листов, начинающихся с 'Раздел'")
    
    # Создаем папку для региона
    region_dir = Path(base_folder) / region_name
    region_dir.mkdir(exist_ok=True, parents=True)
    
    # Загружаем исходный Excel файл
    source_wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Обрабатываем каждый лист
    for sheet in section_sheets:
        # Получаем исходный лист
        source_ws = source_wb[sheet]
        
        # Получаем значения объединенных ячеек
        merged_values = {}
        if source_ws.merged_cells:
            merged_values = process_merged_cells(source_ws)
        
        # Создаем новый файл для листа
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        
        # Копируем данные
        for row in source_ws.rows:
            for cell in row:
                if cell is None:
                    continue
                
                # Проверяем, есть ли значение в объединенных ячейках
                value = merged_values.get((cell.row, cell.column), cell.value)
                # Очищаем значение от переносов строк
                value = clean_value(value)
                
                new_cell = new_ws.cell(
                    row=cell.row,
                    column=cell.column,
                    value=value
                )
        
        # Копируем размеры столбцов
        for col in range(1, source_ws.max_column + 1):
            letter = get_column_letter(col)
            if letter in source_ws.column_dimensions:
                new_ws.column_dimensions[letter].width = source_ws.column_dimensions[letter].width
        
        # Сохраняем результат
        output_path = region_dir / f"{sheet}.xlsx"
        new_wb.save(str(output_path))
        new_wb.close()
    
    source_wb.close()

def process_regional_files(folder_path: str = "2024") -> Dict[str, List[str]]:
    """
    Основная функция для обработки всех региональных Excel файлов.
    
    Args:
        folder_path (str): Путь к папке с региональными файлами (по умолчанию "2024")
        
    Returns:
        Dict[str, List[str]]: Словарь с результатами обработки
            - 'successful': список успешно обработанных регионов
            - 'failed': список регионов с ошибками
    """
    print("=" * 60)
    print("ПАКЕТНАЯ ОБРАБОТКА РЕГИОНАЛЬНЫХ EXCEL ФАЙЛОВ")
    print("=" * 60)
    
    # Получаем список всех Excel файлов
    try:
        excel_files = get_excel_files_from_folder(folder_path)
        print(f"Найдено Excel файлов: {len(excel_files)}")
    except Exception as e:
        print(f"Ошибка при сканировании папки {folder_path}: {str(e)}")
        return {'successful': [], 'failed': []}
    
    if not excel_files:
        print("Excel файлы не найдены!")
        return {'successful': [], 'failed': []}
    
    successful_regions = []
    failed_regions = []
    
    print("\nНачинаем обработку файлов...")
    print("-" * 60)
    
    for i, excel_file in enumerate(excel_files, 1):
        filename = Path(excel_file).name
        region_name = extract_region_name(filename)
        
        print(f"\nОбработка файла {i} из {len(excel_files)}: {region_name}")
        print(f"Файл: {filename}")
        
        try:
            # Проверяем, не существует ли уже папка региона
            region_folder = Path(folder_path) / region_name
            if region_folder.exists() and any(region_folder.iterdir()):
                print(f"  ⚠️  Папка {region_name} уже существует и не пуста. Пропускаем...")
                continue
            
            # Обрабатываем файл
            save_sheets_to_region_folder(excel_file, region_name, folder_path)
            
            # Проверяем результат
            created_files = list(region_folder.glob("*.xlsx")) if region_folder.exists() else []
            print(f"  ✅ Успешно обработан! Создано файлов: {len(created_files)}")
            successful_regions.append(region_name)
            
        except Exception as e:
            print(f"  ❌ Ошибка при обработке: {str(e)}")
            failed_regions.append({
                'region': region_name,
                'file': filename,
                'error': str(e)
            })
            
            # Выводим детальную информацию об ошибке в отладочном режиме
            if os.getenv('DEBUG'):
                print(f"  Детали ошибки:")
                print(f"  {traceback.format_exc()}")
    
    # Выводим финальную сводку
    print("\n" + "=" * 60)
    print("СВОДКА ОБРАБОТКИ")
    print("=" * 60)
    print(f"Всего файлов: {len(excel_files)}")
    print(f"Успешно обработано: {len(successful_regions)}")
    print(f"Ошибок: {len(failed_regions)}")
    
    if successful_regions:
        print(f"\n✅ Успешно обработанные регионы ({len(successful_regions)}):")
        for region in successful_regions:
            print(f"  • {region}")
    
    if failed_regions:
        print(f"\n❌ Регионы с ошибками ({len(failed_regions)}):")
        for failed in failed_regions:
            print(f"  • {failed['region']}: {failed['error']}")
    
    print("\n" + "=" * 60)
    
    return {
        'successful': successful_regions,
        'failed': failed_regions
    }

if __name__ == "__main__":
    # Запуск обработки
    try:
        results = process_regional_files()
        
        # Выход с соответствующим кодом
        if results['failed']:
            print(f"\nОбработка завершена с ошибками. Проверьте {len(results['failed'])} файлов.")
            exit(1)
        else:
            print(f"\nВсе файлы успешно обработаны!")
            exit(0)
            
    except KeyboardInterrupt:
        print("\n\nОбработка прервана пользователем.")
        exit(130)
    except Exception as e:
        print(f"\nКритическая ошибка: {str(e)}")
        if os.getenv('DEBUG'):
            traceback.print_exc()
        exit(1)